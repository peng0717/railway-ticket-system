#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""铁路售票系统数据导出到Excel"""

import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import os

DB_PATH = "./铁路售票系统/data/railway.db"
OUTPUT_PATH = "./铁路售票系统/铁路数据汇总.xlsx"

def get_station_name_map(conn):
    """获取站名映射字典 station_code -> station_name"""
    cursor = conn.execute("SELECT station_code, station_name FROM stations")
    return {row[0]: row[1] for row in cursor.fetchall()}

def get_train_number_map(conn):
    """获取train_id -> train_number映射"""
    cursor = conn.execute("SELECT train_id, train_number FROM trains")
    return {row[0]: row[1] for row in cursor.fetchall()}

def auto_adjust_column_width(ws):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def style_header(ws, headers):
    """设置表头样式"""
    font = Font(bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font
        cell.alignment = alignment

def export_stations(conn, wb):
    """导出车站信息"""
    print("正在导出 Sheet1: 车站信息...")
    ws = wb.create_sheet("车站信息")
    
    headers = ["电报码", "站名", "拼音码"]
    ws.append(headers)
    style_header(ws, headers)
    
    cursor = conn.execute("""
        SELECT telecode, station_name, pinyin_code 
        FROM stations 
        ORDER BY station_id
    """)
    
    for row in cursor:
        ws.append([row[0] or "", row[1] or "", row[2] or ""])
    
    auto_adjust_column_width(ws)
    print(f"  已导出 {cursor.rowcount if cursor.rowcount > 0 else ws.max_row - 1} 条记录")

def export_trains(conn, wb):
    """导出的次信息"""
    print("正在导出 Sheet2: 车次信息...")
    ws = wb.create_sheet("车次信息")
    
    headers = ["车次号", "类型", "始发站", "终到站", "发车时间", "到达时间", "总里程", "运行日期", "状态"]
    ws.append(headers)
    style_header(ws, headers)
    
    cursor = conn.execute("""
        SELECT train_number, train_type, start_station, end_station,
               start_time, end_time, total_distance, running_days, status
        FROM trains 
        ORDER BY train_id
    """)
    
    for row in cursor:
        ws.append(list(row))
    
    auto_adjust_column_width(ws)
    print(f"  已导出 {cursor.rowcount if cursor.rowcount > 0 else ws.max_row - 1} 条记录")

def export_train_stops(conn, wb, train_map, station_map):
    """导出的停站时刻表"""
    print("正在导出 Sheet3: 经停站时刻表...")
    ws = wb.create_sheet("经停站时刻表")
    
    headers = ["车次号", "站序", "站名", "到站时间", "发站时间", "停站时长(分)", "累计里程"]
    ws.append(headers)
    style_header(ws, headers)
    
    cursor = conn.execute("""
        SELECT train_id, stop_sequence, station_code, 
               arrival_time, departure_time, distance_from_start
        FROM train_stops 
        ORDER BY train_id, stop_sequence
    """)
    
    count = 0
    for row in cursor.fetchall():
        train_id, stop_seq, station_code, arrival, departure, distance = row
        
        train_number = train_map.get(train_id, "")
        station_name = station_map.get(station_code, station_code)
        
        # 计算停站时长
        stop_duration = ""
        if arrival and departure:
            try:
                if arrival == departure:
                    stop_duration = 0
                else:
                    t1 = datetime.strptime(arrival, "%H:%M")
                    t2 = datetime.strptime(departure, "%H:%M")
                    diff = (t2 - t1).seconds // 60
                    stop_duration = diff
            except:
                stop_duration = ""
        
        ws.append([train_number, stop_seq, station_name, arrival or "", 
                   departure or "", stop_duration, distance or 0])
        count += 1
    
    auto_adjust_column_width(ws)
    print(f"  已导出 {count} 条记录")

def export_ticket_prices(conn, wb, station_map):
    """导出票价信息"""
    print("正在导出 Sheet4: 票价信息...")
    ws = wb.create_sheet("票价信息")
    
    headers = ["出发站", "到达站", "席别", "票价"]
    ws.append(headers)
    style_header(ws, headers)
    
    cursor = conn.execute("""
        SELECT from_station, to_station, seat_type, base_price
        FROM ticket_prices 
        ORDER BY from_station, to_station, seat_type
    """)
    
    count = 0
    for row in cursor.fetchall():
        from_code, to_code, seat_type, price = row
        from_station = station_map.get(from_code, from_code)
        to_station = station_map.get(to_code, to_code)
        ws.append([from_station, to_station, seat_type, price])
        count += 1
    
    auto_adjust_column_width(ws)
    print(f"  已导出 {count} 条记录")

def main():
    print("=" * 50)
    print("铁路售票系统数据导出工具")
    print("=" * 50)
    
    # 创建输出目录
    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    
    # 连接数据库
    conn = sqlite3.connect(DB_PATH)
    
    # 预加载映射数据
    print("加载映射数据...")
    station_map = get_station_name_map(conn)
    train_map = get_train_number_map(conn)
    print(f"  车站数量: {len(station_map)}")
    print(f"  车次数量: {len(train_map)}")
    
    # 创建工作簿
    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)
    
    # 导出各个Sheet
    export_stations(conn, wb)
    export_trains(conn, wb)
    export_train_stops(conn, wb, train_map, station_map)
    export_ticket_prices(conn, wb, station_map)
    
    # 保存文件
    print("\n正在保存文件...")
    wb.save(OUTPUT_PATH)
    print(f"导出完成！文件已保存至: {OUTPUT_PATH}")
    print("=" * 50)
    
    conn.close()

if __name__ == "__main__":
    main()
